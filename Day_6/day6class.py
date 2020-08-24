# -*- coding: utf-8 -*-
"""
Created on Mon Aug 17 14:55:56 2020

@author: Satish Kumar
"""
import csv
import pymysql

cityset = set()
try:
    file = "realestate.csv"
    with open(file,"r") as fobj:
        reader = csv.reader(fobj)
        
        for line in reader:
            cityset.add(line[1])
        for city in cityset:
            print(city)
except FileNotFoundError as err:
    print("file doesn't exist")
except Exception as err:
    print("Unknown err: ", err)
    
from openpyxl import Workbook
wb = Workbook()

ws = wb.active()
ws['A1']= 42

ws.append([1,2,3])

import datetime
ws['A2'] = datetime.datetime.now()

wb.save("sample.xlsx")