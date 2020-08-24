# -*- coding: utf-8 -*-
"""
Created on Mon Aug 17 16:12:38 2020

@author: gsripath
"""

from openpyxl import Workbook
wb = Workbook()

# grab the active worksheet  # creating one sheet
ws = wb.active

# Data can be assigned directly to cells
ws['A1'] = 42

# Rows can also be appended
ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
ws['A2'] = datetime.datetime.now()

# Save the file
wb.save("sample.xlsx")